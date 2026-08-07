"""Restore the database from a full-workbook backup (the Export page's `database.xlsx`).

Semantics: **make the database match the workbook.** Rows are matched by id (the export
always carries ids); rows missing from the file are deleted, rows missing from the
database are recreated with their original ids, and differing fields are overwritten. A
column absent from the file (an older backup format) leaves that field untouched rather
than nulling it, so old backups restore without losing what they never captured.

Two phases share this one code path: `plan_restore(apply=False)` parses, diffs, and
returns the `RestorePlan` the admin previews; `apply=True` recomputes the same plan from
the same bytes and executes it. Restore is **all-or-nothing** — any parse or reference
error aborts with the errors in the plan and nothing applied (unlike the per-row
best-effort imports in `spreadsheet.py`). One commit at the very end; settings rows are
upserted directly here rather than via `settings.save_settings`, which commits mid-call.

What restore never touches: the admin password hash (deliberately absent from the
export) and anything env-configured (the printer device path).
"""

from __future__ import annotations

import json
import uuid
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Condition,
    Event,
    EventType,
    Group,
    Item,
    ItemType,
    Setting,
    User,
    UserStatus,
)
from app.schemas.restore import (
    DETAIL_CAP,
    EntityPlan,
    FieldChange,
    RestorePlan,
    RowChange,
    SettingChange,
    SheetError,
)
from app.schemas.settings import AppSettings, AppSettingsUpdate
from app.services import settings as settings_svc
from app.services.spreadsheet import _parse_bool, _parse_enum, _parse_uuid, _text

REQUIRED_SHEETS = ("users", "groups", "item_types", "items", "history", "settings")


# ---------------------------------------------------------------------------
# Parsing: workbook bytes -> typed rows, collecting every error
# ---------------------------------------------------------------------------


@dataclass
class _Row:
    id: uuid.UUID
    row: int  # 1-based sheet row, for error messages
    label: str
    fields: dict[str, Any]  # only the columns the file carries, values normalized


@dataclass
class _ParsedEvent:
    id: uuid.UUID
    row: int
    created_at: datetime  # tz-aware UTC
    event_type: EventType
    item_barcode: str | None
    user_barcode: str | None
    note: str | None


@dataclass
class _Parsed:
    groups: dict[uuid.UUID, _Row] = field(default_factory=dict)
    item_types: dict[uuid.UUID, _Row] = field(default_factory=dict)
    users: dict[uuid.UUID, _Row] = field(default_factory=dict)
    items: dict[uuid.UUID, _Row] = field(default_factory=dict)
    events: dict[uuid.UUID, _ParsedEvent] = field(default_factory=dict)
    settings: dict[str, Any] = field(default_factory=dict)  # validated AppSettings keys
    errors: list[SheetError] = field(default_factory=list)


def _read_sheets(content: bytes) -> dict[str, list[tuple[int, dict]]] | None:
    """{sheet_title: [(row_number, {header: value}), ...]} — like spreadsheet._read_rows
    but for every sheet in the workbook."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheets: dict[str, list[tuple[int, dict]]] = {}
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        raw_headers = next(rows, None) or []
        headers = [str(h).strip().lower() if h is not None else "" for h in raw_headers]
        out: list[tuple[int, dict]] = []
        for number, raw in enumerate(rows, start=2):
            record = {headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)}
            if any(v is not None and str(v).strip() != "" for v in record.values()):
                out.append((number, record))
        sheets[ws.title.strip().lower()] = out
    wb.close()
    return sheets


def _parse_date(value: object, sheet: str, row: int, fld: str, errors: list[SheetError]):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        errors.append(SheetError(sheet=sheet, row=row, message=f"Invalid {fld}: {value!r}."))
        return None


def _parse_local_dt(
    value: object, tz: ZoneInfo, sheet: str, row: int, fld: str, errors: list[SheetError]
) -> datetime | None:
    """Exported timestamps are naive local wall time; bring them back to aware UTC."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if not isinstance(value, datetime):
        try:
            value = datetime.fromisoformat(str(value).strip())
        except ValueError:
            errors.append(SheetError(sheet=sheet, row=row, message=f"Invalid {fld}: {value!r}."))
            return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=tz)
    return value.astimezone(UTC)


def _parse_decimal(value: object, sheet: str, row: int, fld: str, errors: list[SheetError]):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        errors.append(SheetError(sheet=sheet, row=row, message=f"Invalid {fld}: {value!r}."))
        return None


def _parse_settings(rows: list[tuple[int, dict]], errors: list[SheetError]) -> dict[str, Any]:
    """Validated AppSettings keys from the key/value sheet; unknown keys are ignored so
    backups survive settings evolving in either direction."""
    raw = {}
    for number, rec in rows:
        key = _text(rec.get("key"))
        if key is None:
            errors.append(SheetError(sheet="settings", row=number, message="Missing key."))
        elif key in AppSettingsUpdate.model_fields:
            raw[key] = rec.get("value")
    try:
        update = AppSettingsUpdate(**raw)
    except ValueError as exc:
        errors.append(SheetError(sheet="settings", message=f"Invalid settings: {exc}"))
        return {}
    return {k: v for k, v in update.model_dump(exclude_unset=True).items()}


def _parse(content: bytes) -> _Parsed:
    parsed = _Parsed()
    try:
        sheets = _read_sheets(content)
    except Exception:
        parsed.errors.append(
            SheetError(
                sheet="workbook",
                message="Not a readable .xlsx workbook. Upload a Stocky database export.",
            )
        )
        return parsed
    for name in REQUIRED_SHEETS:
        if name not in sheets:
            parsed.errors.append(
                SheetError(
                    sheet=name,
                    message="Sheet is missing — this file is not a full-database export.",
                )
            )
    if parsed.errors:
        return parsed

    parsed.settings = _parse_settings(sheets["settings"], parsed.errors)
    tz_name = parsed.settings.get("timezone") or AppSettings().timezone
    tz = ZoneInfo(tz_name)
    errors = parsed.errors

    def rows_of(sheet: str, target: dict[uuid.UUID, _Row], build) -> None:
        for number, rec in sheets[sheet]:
            try:
                row_id = _parse_uuid(rec.get("id"), "id")
            except ValueError as exc:
                errors.append(SheetError(sheet=sheet, row=number, message=str(exc)))
                continue
            if row_id in target:
                errors.append(SheetError(sheet=sheet, row=number, message="Duplicate id."))
                continue
            before = len(errors)
            fields, label = build(number, rec)
            if len(errors) == before:
                target[row_id] = _Row(id=row_id, row=number, label=label, fields=fields)

    def require(rec: dict, key: str, sheet: str, number: int) -> str | None:
        value = _text(rec.get(key))
        if value is None:
            errors.append(SheetError(sheet=sheet, row=number, message=f"Missing {key}."))
        return value

    def build_group(number: int, rec: dict):
        fields: dict[str, Any] = {"name": require(rec, "name", "groups", number)}
        if "parent" in rec:
            fields["parent"] = _text(rec.get("parent"))
        if "semester_start" in rec:
            fields["semester_start"] = _parse_date(
                rec.get("semester_start"), "groups", number, "semester_start", errors
            )
        if "created_at" in rec:
            fields["created_at"] = _parse_local_dt(
                rec.get("created_at"), tz, "groups", number, "created_at", errors
            )
        if "permissions" in rec:
            text = _text(rec.get("permissions"))
            try:
                fields["permissions"] = json.loads(text) if text else {}
            except json.JSONDecodeError:
                errors.append(
                    SheetError(sheet="groups", row=number, message="Invalid permissions JSON.")
                )
        return fields, fields["name"] or "?"

    def build_item_type(number: int, rec: dict):
        fields: dict[str, Any] = {"name": require(rec, "name", "item_types", number)}
        for key in ("manufacturer", "author", "description", "photo_url", "url", "upc_isbn"):
            if key in rec:
                fields[key] = _text(rec.get(key))
        if "publish_date" in rec:
            fields["publish_date"] = _parse_date(
                rec.get("publish_date"), "item_types", number, "publish_date", errors
            )
        if "cost" in rec:
            fields["cost"] = _parse_decimal(rec.get("cost"), "item_types", number, "cost", errors)
        if "created_at" in rec:
            fields["created_at"] = _parse_local_dt(
                rec.get("created_at"), tz, "item_types", number, "created_at", errors
            )
        return fields, fields["name"] or "?"

    def build_user(number: int, rec: dict):
        name = require(rec, "name", "users", number)
        barcode = require(rec, "barcode", "users", number)
        fields: dict[str, Any] = {"name": name, "barcode": barcode}
        if "group" in rec:
            fields["group"] = _text(rec.get("group"))
        try:
            fields["status"] = str(_parse_enum(UserStatus, rec.get("status") or "Active", "status"))
        except ValueError as exc:
            errors.append(SheetError(sheet="users", row=number, message=str(exc)))
        if "created_at" in rec:
            fields["created_at"] = _parse_local_dt(
                rec.get("created_at"), tz, "users", number, "created_at", errors
            )
        return fields, f"{name or '?'} ({barcode or '?'})"

    def build_item(number: int, rec: dict):
        name = require(rec, "name", "items", number)
        barcode = require(rec, "barcode", "items", number)
        fields: dict[str, Any] = {
            "name": name,
            "barcode": barcode,
            "item_type": require(rec, "item_type", "items", number),
        }
        if "location" in rec:
            fields["location"] = _text(rec.get("location"))
        try:
            fields["condition"] = str(
                _parse_enum(Condition, rec.get("condition") or "New", "condition")
            )
        except ValueError as exc:
            errors.append(SheetError(sheet="items", row=number, message=str(exc)))
        if "needs_review" in rec:
            raw = rec.get("needs_review")
            fields["needs_review"] = raw if isinstance(raw, bool) else _parse_bool(raw)
        for key in ("photo_url", "description"):
            if key in rec:
                fields[key] = _text(rec.get(key))
        if "purchase_price" in rec:
            fields["purchase_price"] = _parse_decimal(
                rec.get("purchase_price"), "items", number, "purchase_price", errors
            )
        if "purchase_date" in rec:
            fields["purchase_date"] = _parse_date(
                rec.get("purchase_date"), "items", number, "purchase_date", errors
            )
        if "created_at" in rec:
            fields["created_at"] = _parse_local_dt(
                rec.get("created_at"), tz, "items", number, "created_at", errors
            )
        return fields, f"{name or '?'} ({barcode or '?'})"

    rows_of("groups", parsed.groups, build_group)
    rows_of("item_types", parsed.item_types, build_item_type)
    rows_of("users", parsed.users, build_user)
    rows_of("items", parsed.items, build_item)

    # History: resolve references by barcode within the *file*, so a consistent backup
    # can be validated without touching the database.
    items_by_barcode = {r.fields["barcode"]: r for r in parsed.items.values()}
    users_by_barcode = {r.fields["barcode"]: r for r in parsed.users.values()}
    for number, rec in sheets["history"]:
        try:
            event_id = _parse_uuid(rec.get("id"), "id")
            event_type = _parse_enum(EventType, rec.get("event_type"), "event_type")
        except ValueError as exc:
            errors.append(SheetError(sheet="history", row=number, message=str(exc)))
            continue
        if event_id in parsed.events:
            errors.append(SheetError(sheet="history", row=number, message="Duplicate id."))
            continue
        created_at = _parse_local_dt(
            rec.get("created_at"), tz, "history", number, "created_at", errors
        )
        if created_at is None:
            errors.append(SheetError(sheet="history", row=number, message="Missing created_at."))
            continue
        item_barcode = _text(rec.get("item_barcode"))
        user_barcode = _text(rec.get("user_barcode"))
        if item_barcode is not None and item_barcode not in items_by_barcode:
            errors.append(
                SheetError(
                    sheet="history",
                    row=number,
                    message=f"item_barcode {item_barcode!r} is not on the items sheet.",
                )
            )
            continue
        if user_barcode is not None and user_barcode not in users_by_barcode:
            errors.append(
                SheetError(
                    sheet="history",
                    row=number,
                    message=f"user_barcode {user_barcode!r} is not on the users sheet.",
                )
            )
            continue
        if item_barcode is None and event_type != EventType.ATTENDANCE:
            errors.append(
                SheetError(
                    sheet="history",
                    row=number,
                    message=f"A {event_type} event needs an item_barcode.",
                )
            )
            continue
        parsed.events[event_id] = _ParsedEvent(
            id=event_id,
            row=number,
            created_at=created_at,
            event_type=event_type,
            item_barcode=item_barcode,
            user_barcode=user_barcode,
            note=_text(rec.get("note")),
        )

    # Names are how sheets reference each other, so they must be unambiguous per file.
    _check_reference_names(parsed)
    _check_unique_barcodes(parsed)
    return parsed


def _check_reference_names(parsed: _Parsed) -> None:
    group_name_counts: dict[str, int] = {}
    for row in parsed.groups.values():
        group_name_counts[row.fields["name"]] = group_name_counts.get(row.fields["name"], 0) + 1
    type_name_counts: dict[str, int] = {}
    for row in parsed.item_types.values():
        type_name_counts[row.fields["name"]] = type_name_counts.get(row.fields["name"], 0) + 1

    def check(row: _Row, sheet: str, fld: str, counts: dict[str, int], target: str) -> None:
        name = row.fields.get(fld)
        if name is None:
            return
        if counts.get(name, 0) == 0:
            parsed.errors.append(
                SheetError(
                    sheet=sheet, row=row.row, message=f"Unknown {target} {name!r} for {fld}."
                )
            )
        elif counts[name] > 1:
            parsed.errors.append(
                SheetError(
                    sheet=sheet,
                    row=row.row,
                    message=f"{target} name {name!r} is ambiguous in this file.",
                )
            )

    for row in parsed.groups.values():
        check(row, "groups", "parent", group_name_counts, "group")
    for row in parsed.users.values():
        check(row, "users", "group", group_name_counts, "group")
    for row in parsed.items.values():
        check(row, "items", "item_type", type_name_counts, "item type")


def _check_unique_barcodes(parsed: _Parsed) -> None:
    for sheet, rows in (("users", parsed.users), ("items", parsed.items)):
        seen: dict[str, int] = {}
        for row in rows.values():
            code = row.fields.get("barcode")
            if code in seen:
                parsed.errors.append(
                    SheetError(
                        sheet=sheet,
                        row=row.row,
                        message=f"Barcode {code!r} appears more than once.",
                    )
                )
            seen[code] = row.row


# ---------------------------------------------------------------------------
# Diff: parsed file vs live database
# ---------------------------------------------------------------------------


def _values_equal(a: Any, b: Any) -> bool:
    if isinstance(a, datetime) and isinstance(b, datetime):
        return abs((a - b).total_seconds()) < 1  # xlsx serial time wobbles below 1s
    if isinstance(a, Decimal) and isinstance(b, Decimal):
        return a == b
    return a == b


def _display(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, dict):
        return json.dumps(value, sort_keys=True)
    return str(value)


def _diff_entity(
    kind: str,
    file_rows: dict[uuid.UUID, _Row],
    db_objs: dict[uuid.UUID, Any],
    extract,
    db_label,
) -> EntityPlan:
    plan = EntityPlan(kind=kind)
    for row_id, row in file_rows.items():
        obj = db_objs.get(row_id)
        if obj is None:
            plan.create_count += 1
            if len(plan.creates) < DETAIL_CAP:
                plan.creates.append(RowChange(id=row_id, label=row.label))
            continue
        changes = [
            FieldChange(field=fld, old=_display(current), new=_display(wanted))
            for fld, wanted in row.fields.items()
            if not _values_equal((current := extract(obj, fld)), wanted)
        ]
        if changes:
            plan.update_count += 1
            if len(plan.updates) < DETAIL_CAP:
                plan.updates.append(RowChange(id=row_id, label=row.label, fields=changes))
        else:
            plan.unchanged += 1
    for obj_id, obj in db_objs.items():
        if obj_id not in file_rows:
            plan.delete_count += 1
            if len(plan.deletes) < DETAIL_CAP:
                plan.deletes.append(RowChange(id=obj_id, label=db_label(obj)))
    plan.truncated = (
        plan.create_count > len(plan.creates)
        or plan.update_count > len(plan.updates)
        or plan.delete_count > len(plan.deletes)
    )
    return plan


@dataclass
class _DbState:
    groups: dict[uuid.UUID, Group]
    item_types: dict[uuid.UUID, ItemType]
    users: dict[uuid.UUID, User]
    items: dict[uuid.UUID, Item]
    events: dict[uuid.UUID, Event]


async def _load_db(session: AsyncSession) -> _DbState:
    return _DbState(
        groups={g.id: g for g in (await session.execute(select(Group))).scalars()},
        item_types={t.id: t for t in (await session.execute(select(ItemType))).scalars()},
        users={u.id: u for u in (await session.execute(select(User))).scalars()},
        items={i.id: i for i in (await session.execute(select(Item))).scalars()},
        events={e.id: e for e in (await session.execute(select(Event))).scalars()},
    )


def _event_targets(parsed: _Parsed):
    """event id -> (item_id, user_id), resolved once via the file's barcode maps."""
    items_by_barcode = {r.fields["barcode"]: r.id for r in parsed.items.values()}
    users_by_barcode = {r.fields["barcode"]: r.id for r in parsed.users.values()}
    return {
        e.id: (
            items_by_barcode.get(e.item_barcode) if e.item_barcode else None,
            users_by_barcode.get(e.user_barcode) if e.user_barcode else None,
        )
        for e in parsed.events.values()
    }


def _build_plan(parsed: _Parsed, db: _DbState, current_settings: AppSettings) -> RestorePlan:
    group_names_db = {g.id: g.name for g in db.groups.values()}
    type_names_db = {t.id: t.name for t in db.item_types.values()}

    def extract_group(g: Group, fld: str) -> Any:
        match fld:
            case "parent":
                return group_names_db.get(g.parent_id) if g.parent_id else None
            case "permissions":
                return g.permissions or {}
            case _:
                return getattr(g, fld)

    def extract_user(u: User, fld: str) -> Any:
        match fld:
            case "group":
                return group_names_db.get(u.group_id) if u.group_id else None
            case "status":
                return str(u.status)
            case _:
                return getattr(u, fld)

    def extract_item(i: Item, fld: str) -> Any:
        match fld:
            case "item_type":
                return type_names_db.get(i.item_type_id)
            case "condition":
                return str(i.condition)
            case _:
                return getattr(i, fld)

    def extract_type(t: ItemType, fld: str) -> Any:
        return getattr(t, fld)

    plan = RestorePlan()
    plan.entities = [
        _diff_entity("groups", parsed.groups, db.groups, extract_group, lambda g: g.name),
        _diff_entity(
            "item_types", parsed.item_types, db.item_types, extract_type, lambda t: t.name
        ),
        _diff_entity(
            "users", parsed.users, db.users, extract_user, lambda u: f"{u.name} ({u.barcode})"
        ),
        _diff_entity(
            "items", parsed.items, db.items, extract_item, lambda i: f"{i.name} ({i.barcode})"
        ),
    ]

    targets = _event_targets(parsed)
    for event_id, event in parsed.events.items():
        existing = db.events.get(event_id)
        if existing is None:
            plan.events_add += 1
            continue
        item_id, user_id = targets[event_id]
        same = (
            existing.item_id == item_id
            and existing.user_id == user_id
            and existing.event_type == event.event_type
            and (existing.note or None) == event.note
            and _values_equal(existing.created_at, event.created_at)
        )
        if same:
            plan.events_unchanged += 1
        else:
            plan.events_relink += 1
    plan.events_remove = sum(1 for e in db.events if e not in parsed.events)

    current = current_settings.model_dump()
    for key, wanted in parsed.settings.items():
        if current.get(key) != wanted:
            old, new = _display(current.get(key)) or "", _display(wanted) or ""
            plan.settings.append(SettingChange(key=key, old=old, new=new))
    return plan


# ---------------------------------------------------------------------------
# Apply: one transaction, committed only at the end
# ---------------------------------------------------------------------------


async def _apply(session: AsyncSession, parsed: _Parsed, db: _DbState) -> None:
    file_event_ids = set(parsed.events)

    # 1. Events not in the backup go first: they may reference items/users that are
    #    themselves about to be deleted.
    doomed_events = [e for e in db.events if e not in file_event_ids]
    if doomed_events:
        await session.execute(delete(Event).where(Event.id.in_(doomed_events)))

    # 2. Item types, then groups (parents resolved in a second pass, since a parent row
    #    may itself be new).
    for row in parsed.item_types.values():
        obj = db.item_types.get(row.id) or ItemType(id=row.id, name=row.fields["name"])
        for fld, value in row.fields.items():
            setattr(obj, fld, value)
        session.add(obj)
    for row in parsed.groups.values():
        obj = db.groups.get(row.id) or Group(id=row.id, name=row.fields["name"])
        for fld, value in row.fields.items():
            if fld != "parent":
                setattr(obj, fld, value)
        session.add(obj)
    await session.flush()
    groups_by_name = {r.fields["name"]: r.id for r in parsed.groups.values()}
    for row in parsed.groups.values():
        if "parent" in row.fields:
            parent_name = row.fields["parent"]
            obj = db.groups.get(row.id) or await session.get(Group, row.id)
            obj.parent_id = groups_by_name[parent_name] if parent_name else None
            session.add(obj)
    types_by_name = {r.fields["name"]: r.id for r in parsed.item_types.values()}

    # 3. Barcodes are unique; when the backup moves one between rows the intermediate
    #    state would collide. Park every conflicting live barcode first.
    for objs, rows in ((db.users, parsed.users), (db.items, parsed.items)):
        wanted = {r.fields["barcode"] for r in rows.values()}
        for obj in objs.values():
            target = rows.get(obj.id)
            keeps_own = target is not None and target.fields["barcode"] == obj.barcode
            if not keeps_own and obj.barcode in wanted:
                obj.barcode = f"~restore-{obj.id.hex}"
                session.add(obj)
    await session.flush()

    # 4. Users and items.
    for row in parsed.users.values():
        obj = db.users.get(row.id) or User(
            id=row.id, name=row.fields["name"], barcode=row.fields["barcode"]
        )
        for fld, value in row.fields.items():
            if fld == "group":
                obj.group_id = groups_by_name[value] if value else None
            else:
                setattr(obj, fld, value)
        session.add(obj)
    for row in parsed.items.values():
        obj = db.items.get(row.id) or Item(
            id=row.id,
            name=row.fields["name"],
            barcode=row.fields["barcode"],
            item_type_id=types_by_name[row.fields["item_type"]],
        )
        for fld, value in row.fields.items():
            if fld == "item_type":
                obj.item_type_id = types_by_name[value]
            else:
                setattr(obj, fld, value)
        session.add(obj)
    await session.flush()

    # 5. Events: correct kept rows that drifted (hand-edited backups, anonymization) and
    #    reinsert the ones the backup has but the database lost.
    targets = _event_targets(parsed)
    for event_id, event in parsed.events.items():
        item_id, user_id = targets[event_id]
        existing = db.events.get(event_id)
        if existing is None:
            session.add(
                Event(
                    id=event.id,
                    item_id=item_id,
                    user_id=user_id,
                    event_type=event.event_type,
                    note=event.note,
                    created_at=event.created_at,
                )
            )
        else:
            existing.item_id = item_id
            existing.user_id = user_id
            existing.event_type = event.event_type
            existing.note = event.note
            if not _values_equal(existing.created_at, event.created_at):
                existing.created_at = event.created_at
            session.add(existing)
    await session.flush()

    # 6. Deletions, leaves first: items and users, then groups (children before
    #    parents), then item types.
    for item_id in set(db.items) - set(parsed.items):
        await session.delete(db.items[item_id])
    for user_id in set(db.users) - set(parsed.users):
        await session.delete(db.users[user_id])
    await session.flush()
    doomed_groups = {gid: g for gid, g in db.groups.items() if gid not in parsed.groups}
    while doomed_groups:
        parents_of_doomed = {g.parent_id for g in doomed_groups.values()}
        leaves = [gid for gid in doomed_groups if gid not in parents_of_doomed]
        for gid in leaves:
            await session.delete(doomed_groups.pop(gid))
        await session.flush()
    for type_id in set(db.item_types) - set(parsed.item_types):
        await session.delete(db.item_types[type_id])
    await session.flush()

    # 7. Settings — direct row upserts; settings.save_settings would commit mid-restore.
    for key, value in parsed.settings.items():
        row = await session.get(Setting, key)
        if row is None:
            session.add(Setting(key=key, value=value))
        else:
            row.value = value
            session.add(row)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


async def plan_restore(session: AsyncSession, content: bytes, *, apply: bool) -> RestorePlan:
    """Diff the uploaded backup against the database; execute the diff when `apply`.

    Never partially applies: errors abort before any write, and the whole apply is a
    single transaction committed at the end.
    """
    parsed = _parse(content)
    if parsed.errors:
        return RestorePlan(errors=parsed.errors)
    db = await _load_db(session)
    plan = _build_plan(parsed, db, await settings_svc.load_settings(session))
    if apply:
        await _apply(session, parsed, db)
        await session.commit()
        plan.applied = True
    return plan
