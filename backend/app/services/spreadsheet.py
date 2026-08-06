"""XLSX export/import of users and items.

Each sheet leads with an `action` column (C/U/D/blank) so an admin can round-trip the data
through Excel: download, edit, set actions, and re-upload. Import is best-effort — every row is
attempted and a per-row error never aborts the others; the response summarizes what happened.

openpyxl is pure-Python (Pi-friendly); we stream with write_only/read_only to keep memory low.
"""

from __future__ import annotations

import uuid
from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Condition, Event, EventType, Group, Item, ItemType, User, UserStatus
from app.schemas.imports import ImportResult, RowError
from app.services import barcode as barcode_svc
from app.services import events as events_svc
from app.services import settings as settings_svc
from app.services.queries import event_filter_query, group_names, item_type_names

USER_HEADERS = ["action", "id", "barcode", "name", "group", "status"]
ITEM_HEADERS = [
    "action",
    "id",
    "barcode",
    "name",
    "item_type",
    "location",
    "condition",
    "needs_review",
]
# History is export-only — no action column, so it can't be mistaken for an importable sheet.
EVENT_HEADERS = [
    "id",
    "created_at",
    "event_type",
    "item",
    "item_barcode",
    "user",
    "user_barcode",
    "note",
]
# Export-only sheets of the full-database workbook.
GROUP_HEADERS = ["id", "name", "parent", "created_at"]
ITEM_TYPE_HEADERS = [
    "id",
    "name",
    "manufacturer",
    "author",
    "publish_date",
    "description",
    "photo_url",
    "url",
    "cost",
    "upc_isbn",
]


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def _workbook_bytes(headers: list[str], rows: list[list]) -> bytes:
    return _multi_sheet_bytes([(None, headers, rows)])


def _multi_sheet_bytes(sheets: list[tuple[str | None, list[str], list[list]]]) -> bytes:
    wb = Workbook(write_only=True)
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for row in rows:
            ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _read_rows(content: bytes) -> list[tuple[int, dict]]:
    """Yield (row_number, {header: value}) for each data row, header-keyed and lower-cased."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    raw_headers = next(rows, None) or []
    headers = [str(h).strip().lower() if h is not None else "" for h in raw_headers]
    out: list[tuple[int, dict]] = []
    for number, raw in enumerate(rows, start=2):  # row 1 is the header
        record = {headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)}
        out.append((number, record))
    wb.close()
    return out


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_bool(value: object) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def _parse_uuid(value: object, field: str) -> uuid.UUID:
    try:
        return uuid.UUID(str(value).strip())
    except (ValueError, AttributeError) as exc:
        raise ValueError(f"Invalid {field}: {value!r}.") from exc


def _parse_enum[E](enum_cls: type[E], value: object, field: str) -> E:
    text = _text(value)
    try:
        return enum_cls(text)
    except ValueError as exc:
        allowed = ", ".join(e.value for e in enum_cls)  # type: ignore[attr-defined]
        raise ValueError(f"Invalid {field} {text!r} (allowed: {allowed}).") from exc


# ---------------------------------------------------------------------------
# Users
# ---------------------------------------------------------------------------
async def _user_rows(session: AsyncSession) -> list[list]:
    groups = await group_names(session)
    users = (await session.execute(select(User).order_by(User.name))).scalars()
    return [
        ["", str(u.id), u.barcode, u.name, groups.get(u.group_id, ""), str(u.status)] for u in users
    ]


async def users_workbook(session: AsyncSession) -> bytes:
    return _workbook_bytes(USER_HEADERS, await _user_rows(session))


async def import_users(session: AsyncSession, content: bytes) -> ImportResult:
    result = ImportResult()
    groups_by_name = {g.name: g for g in (await session.execute(select(Group))).scalars()}

    async def find(rec: dict) -> User | None:
        if _text(rec.get("id")):
            return await session.get(User, _parse_uuid(rec["id"], "id"))
        code = _text(rec.get("barcode"))
        if code:
            return (
                await session.execute(select(User).where(User.barcode == code))
            ).scalar_one_or_none()
        raise ValueError("Need an id or barcode to match a user.")

    def resolve_group(rec: dict) -> uuid.UUID | None:
        name = _text(rec.get("group"))
        if name is None:
            return None
        group = groups_by_name.get(name)
        if group is None:
            raise ValueError(f"Unknown group {name!r}.")
        return group.id

    for number, rec in _read_rows(content):
        action = (_text(rec.get("action")) or "").upper()
        try:
            if not action:
                result.skipped += 1
            elif action == "C":
                name = _text(rec.get("name"))
                if not name:
                    raise ValueError("Name is required to create a user.")
                user = User(
                    name=name,
                    group_id=resolve_group(rec),
                    status=_parse_enum(UserStatus, rec.get("status") or "Active", "status"),
                    barcode=await barcode_svc.allocate_barcode(
                        session, User, barcode_svc.USER_PREFIX, _text(rec.get("barcode"))
                    ),
                )
                session.add(user)
                await session.flush()
                result.created += 1
            elif action == "U":
                user = await find(rec)
                if user is None:
                    raise ValueError("No matching user to update.")
                if _text(rec.get("name")):
                    user.name = _text(rec["name"])
                if "group" in rec:
                    user.group_id = resolve_group(rec)
                if _text(rec.get("status")):
                    user.status = _parse_enum(UserStatus, rec["status"], "status")
                if _text(rec.get("barcode")):
                    user.barcode = await barcode_svc.allocate_barcode(
                        session, User, barcode_svc.USER_PREFIX, _text(rec["barcode"]), user.barcode
                    )
                session.add(user)
                result.updated += 1
            elif action == "D":
                user = await find(rec)
                if user is None:
                    raise ValueError("No matching user to delete.")
                await events_svc.detach_user_history(session, [user.id])
                await session.delete(user)
                result.deleted += 1
            else:
                raise ValueError(f"Unknown action {action!r} (use C, U, or D).")
        except (ValueError, barcode_svc.BarcodeConflict) as exc:
            result.errors.append(RowError(row=number, message=str(exc)))

    await session.commit()
    return result


# ---------------------------------------------------------------------------
# History (export only)
# ---------------------------------------------------------------------------
async def _event_rows(
    session: AsyncSession,
    tz: ZoneInfo,
    *,
    event_type: EventType | None = None,
    user_id: uuid.UUID | None = None,
    item_id: uuid.UUID | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    q: str | None = None,
) -> list[list]:
    events = (
        await session.execute(
            event_filter_query(
                event_type=event_type,
                user_id=user_id,
                item_id=item_id,
                date_from=date_from,
                date_to=date_to,
                q=q,
            )
        )
    ).all()
    # Barcodes aren't carried by the shared history query; resolve them in two dict lookups
    # rather than widening a query the paginated admin view also uses.
    item_barcodes = dict((await session.execute(select(Item.id, Item.barcode))).all())
    user_barcodes = dict((await session.execute(select(User.id, User.barcode))).all())
    return [
        [
            str(event.id),
            # openpyxl can't write tz-aware datetimes; store the local wall time instead.
            event.created_at.astimezone(tz).replace(tzinfo=None),
            str(event.event_type),
            item_name or "",
            item_barcodes.get(event.item_id, "") if event.item_id else "",
            user_name or "",
            user_barcodes.get(event.user_id, "") if event.user_id else "",
            event.note or "",
        ]
        for event, item_name, user_name in events
    ]


async def events_workbook(
    session: AsyncSession,
    *,
    event_type: EventType | None = None,
    user_id: uuid.UUID | None = None,
    item_id: uuid.UUID | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    q: str | None = None,
) -> bytes:
    """Event history matching the given filters, most recent first."""
    tz = ZoneInfo(await settings_svc.app_timezone(session))
    rows = await _event_rows(
        session,
        tz,
        event_type=event_type,
        user_id=user_id,
        item_id=item_id,
        date_from=date_from,
        date_to=date_to,
        q=q,
    )
    return _workbook_bytes(EVENT_HEADERS, rows)


# ---------------------------------------------------------------------------
# Items
# ---------------------------------------------------------------------------
async def _item_rows(session: AsyncSession) -> list[list]:
    types = await item_type_names(session)
    items = (await session.execute(select(Item).order_by(Item.name))).scalars()
    return [
        [
            "",
            str(i.id),
            i.barcode,
            i.name,
            types.get(i.item_type_id, ""),
            i.location or "",
            str(i.condition),
            i.needs_review,
        ]
        for i in items
    ]


async def items_workbook(session: AsyncSession) -> bytes:
    return _workbook_bytes(ITEM_HEADERS, await _item_rows(session))


async def import_items(session: AsyncSession, content: bytes) -> ImportResult:
    result = ImportResult()
    types_by_name = {t.name: t for t in (await session.execute(select(ItemType))).scalars()}

    async def find(rec: dict) -> Item | None:
        if _text(rec.get("id")):
            return await session.get(Item, _parse_uuid(rec["id"], "id"))
        code = _text(rec.get("barcode"))
        if code:
            return (
                await session.execute(select(Item).where(Item.barcode == code))
            ).scalar_one_or_none()
        raise ValueError("Need an id or barcode to match an item.")

    def resolve_type(rec: dict) -> uuid.UUID:
        name = _text(rec.get("item_type"))
        if name is None:
            raise ValueError("item_type is required.")
        item_type = types_by_name.get(name)
        if item_type is None:
            raise ValueError(f"Unknown item type {name!r}.")
        return item_type.id

    for number, rec in _read_rows(content):
        action = (_text(rec.get("action")) or "").upper()
        try:
            if not action:
                result.skipped += 1
            elif action == "C":
                name = _text(rec.get("name"))
                if not name:
                    raise ValueError("Name is required to create an item.")
                item = Item(
                    name=name,
                    item_type_id=resolve_type(rec),
                    location=_text(rec.get("location")),
                    condition=_parse_enum(Condition, rec.get("condition") or "New", "condition"),
                    needs_review=_parse_bool(rec.get("needs_review")),
                    barcode=await barcode_svc.allocate_barcode(
                        session, Item, barcode_svc.ITEM_PREFIX, _text(rec.get("barcode"))
                    ),
                )
                session.add(item)
                await session.flush()
                session.add(Event(item_id=item.id, event_type=EventType.CREATE))
                result.created += 1
            elif action == "U":
                item = await find(rec)
                if item is None:
                    raise ValueError("No matching item to update.")
                if _text(rec.get("name")):
                    item.name = _text(rec["name"])
                if _text(rec.get("item_type")):
                    item.item_type_id = resolve_type(rec)
                if "location" in rec:
                    item.location = _text(rec.get("location"))
                if _text(rec.get("condition")):
                    item.condition = _parse_enum(Condition, rec["condition"], "condition")
                if _text(rec.get("needs_review")) is not None:
                    item.needs_review = _parse_bool(rec.get("needs_review"))
                if _text(rec.get("barcode")):
                    item.barcode = await barcode_svc.allocate_barcode(
                        session, Item, barcode_svc.ITEM_PREFIX, _text(rec["barcode"]), item.barcode
                    )
                session.add(item)
                result.updated += 1
            elif action == "D":
                item = await find(rec)
                if item is None:
                    raise ValueError("No matching item to delete.")
                await session.execute(delete(Event).where(Event.item_id == item.id))
                await session.delete(item)
                result.deleted += 1
            else:
                raise ValueError(f"Unknown action {action!r} (use C, U, or D).")
        except (ValueError, barcode_svc.BarcodeConflict) as exc:
            result.errors.append(RowError(row=number, message=str(exc)))

    await session.commit()
    return result


# ---------------------------------------------------------------------------
# Full-database export (export only)
# ---------------------------------------------------------------------------
async def full_workbook(session: AsyncSession) -> bytes:
    """Every table as its own sheet, in one workbook.

    The users/items sheets reuse the import format (with a blank action column) so rows can be
    pasted straight into an import file. The settings sheet is built from `load_settings()` —
    i.e. only the keys declared on AppSettings — never from raw Setting rows, so values stored
    outside the schema (the admin password hash lives in the same table) can never appear.
    """
    tz = ZoneInfo(await settings_svc.app_timezone(session))

    groups = list((await session.execute(select(Group).order_by(Group.name))).scalars())
    names_by_id = {g.id: g.name for g in groups}
    group_rows = [
        [
            str(g.id),
            g.name,
            names_by_id.get(g.parent_id, "") if g.parent_id else "",
            g.created_at.astimezone(tz).replace(tzinfo=None),
        ]
        for g in groups
    ]

    types = (await session.execute(select(ItemType).order_by(ItemType.name))).scalars()
    type_rows = [
        [
            str(t.id),
            t.name,
            t.manufacturer or "",
            t.author or "",
            t.publish_date,
            t.description or "",
            t.photo_url or "",
            t.url or "",
            t.cost,
            t.upc_isbn or "",
        ]
        for t in types
    ]

    settings_doc = (await settings_svc.load_settings(session)).model_dump()

    return _multi_sheet_bytes(
        [
            ("users", USER_HEADERS, await _user_rows(session)),
            ("groups", GROUP_HEADERS, group_rows),
            ("item_types", ITEM_TYPE_HEADERS, type_rows),
            ("items", ITEM_HEADERS, await _item_rows(session)),
            ("history", EVENT_HEADERS, await _event_rows(session, tz)),
            ("settings", ["key", "value"], [[k, v] for k, v in settings_doc.items()]),
        ]
    )
