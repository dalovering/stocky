"""Attendance events: recorded on the first kiosk scan of the (local) day."""

from __future__ import annotations

import uuid
from datetime import UTC, datetime, timedelta

import pytest
from sqlalchemy import select

from app.models import Event, EventType


@pytest.fixture
async def user(admin_client):
    return (await admin_client.post("/api/admin/users", json={"name": "Ada"})).json()


async def _attendance_events(session, user_id: str) -> list[Event]:
    return list(
        (
            await session.execute(
                select(Event).where(
                    Event.user_id == uuid.UUID(user_id),
                    Event.event_type == EventType.ATTENDANCE,
                )
            )
        ).scalars()
    )


@pytest.mark.asyncio
async def test_first_scan_of_day_records_attendance_once(admin_client, session, user):
    # First scan logs in AND records attendance.
    scan = (await admin_client.post("/api/kiosk/scan", json={"barcode": user["barcode"]})).json()
    assert scan["action"] == "login"
    assert len(await _attendance_events(session, user["id"])) == 1

    # A second scan the same day logs in but does not duplicate the attendance row.
    scan = (await admin_client.post("/api/kiosk/scan", json={"barcode": user["barcode"]})).json()
    assert scan["action"] == "login"
    assert len(await _attendance_events(session, user["id"])) == 1


@pytest.mark.asyncio
async def test_attendance_appears_in_histories(admin_client, session, user):
    await admin_client.post("/api/kiosk/scan", json={"barcode": user["barcode"]})

    # Kiosk user history (filters purely on user_id).
    kiosk_events = (await admin_client.get(f"/api/kiosk/user/{user['id']}/events")).json()
    assert [e["event_type"] for e in kiosk_events] == ["attendance"]
    assert kiosk_events[0]["item_id"] is None

    # Admin history — regression for the Item join: an inner join would drop null-item rows.
    page = (await admin_client.get("/api/admin/events?event_type=attendance")).json()
    assert page["total"] == 1
    assert page["items"][0]["user_name"] == "Ada"
    assert page["items"][0]["item_id"] is None
    assert page["items"][0]["item_name"] is None


@pytest.mark.asyncio
async def test_blocked_inactive_user_gets_no_attendance(admin_client, session):
    inactive = (
        await admin_client.post("/api/admin/users", json={"name": "Bob", "status": "Inactive"})
    ).json()
    await admin_client.patch("/api/admin/settings", json={"kiosk_block_inactive_users": True})

    scan = (
        await admin_client.post("/api/kiosk/scan", json={"barcode": inactive["barcode"]})
    ).json()
    assert scan["action"] == "unknown"
    assert await _attendance_events(session, inactive["id"]) == []


@pytest.mark.asyncio
async def test_attendance_day_boundary_uses_app_timezone(admin_client, session, user):
    """A scan yesterday (local) doesn't count for today; one today (local) does."""
    # America/New_York is the default timezone setting; build timestamps relative to local days.
    from zoneinfo import ZoneInfo

    tz = ZoneInfo("America/New_York")
    now_local = datetime.now(UTC).astimezone(tz)

    # Attendance recorded yesterday evening local time -> today's scan records a fresh one.
    yesterday = (now_local - timedelta(days=1)).replace(hour=21, minute=0)
    session.add(
        Event(
            item_id=None,
            user_id=uuid.UUID(user["id"]),
            event_type=EventType.ATTENDANCE,
            created_at=yesterday.astimezone(UTC),
        )
    )
    await session.commit()

    await admin_client.post("/api/kiosk/scan", json={"barcode": user["barcode"]})
    assert len(await _attendance_events(session, user["id"])) == 2

    # Scanning again the same local day adds nothing.
    await admin_client.post("/api/kiosk/scan", json={"barcode": user["barcode"]})
    assert len(await _attendance_events(session, user["id"])) == 2


@pytest.mark.asyncio
async def test_attendance_history_export_includes_null_item_rows(admin_client, user):
    """The xlsx export tolerates attendance rows (blank item columns)."""
    from io import BytesIO

    from openpyxl import load_workbook

    await admin_client.post("/api/kiosk/scan", json={"barcode": user["barcode"]})
    content = (await admin_client.get("/api/admin/events.xlsx")).content
    ws = load_workbook(BytesIO(content)).active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    record = dict(zip(headers, rows[1], strict=False))
    assert record["event_type"] == "attendance"
    assert record["item"] in (None, "")
    assert record["user"] == "Ada"


# ---------------------------------------------------------------------------
# Deletion paths: attendance rows go with the user; item history is anonymized.
# ---------------------------------------------------------------------------
async def _seed_history(admin_client, user):
    """Give the user attendance plus a checkout so both event kinds exist."""
    item_type = (await admin_client.post("/api/admin/item-types", json={"name": "Calc"})).json()
    item = (
        await admin_client.post(
            "/api/admin/items", json={"item_type_id": item_type["id"], "name": "Calc #1"}
        )
    ).json()
    await admin_client.post("/api/kiosk/scan", json={"barcode": user["barcode"]})  # attendance
    await admin_client.post(
        "/api/kiosk/checkout", json={"item_id": item["id"], "user_id": user["id"]}
    )
    return item


async def _assert_history_detached(admin_client, session, user):
    assert await _attendance_events(session, user["id"]) == []
    remaining = list(
        (
            await session.execute(select(Event).where(Event.user_id == uuid.UUID(user["id"])))
        ).scalars()
    )
    assert remaining == []  # anonymized, not deleted:
    checkouts = list(
        (
            await session.execute(select(Event).where(Event.event_type == EventType.CHECKOUT))
        ).scalars()
    )
    assert len(checkouts) == 1
    assert checkouts[0].user_id is None


@pytest.mark.asyncio
async def test_single_delete_removes_attendance_and_detaches_history(admin_client, session, user):
    await _seed_history(admin_client, user)
    resp = await admin_client.delete(f"/api/admin/users/{user['id']}")
    assert resp.status_code in (200, 204)
    await _assert_history_detached(admin_client, session, user)


@pytest.mark.asyncio
async def test_batch_delete_removes_attendance_and_detaches_history(admin_client, session, user):
    await _seed_history(admin_client, user)
    resp = await admin_client.post("/api/admin/users/batch-delete", json={"ids": [user["id"]]})
    assert resp.status_code in (200, 204)
    await _assert_history_detached(admin_client, session, user)


@pytest.mark.asyncio
async def test_import_delete_removes_attendance_and_detaches_history(admin_client, session, user):
    from io import BytesIO

    from openpyxl import Workbook

    await _seed_history(admin_client, user)
    wb = Workbook()
    ws = wb.active
    ws.append(["action", "id", "barcode", "name", "group", "status"])
    ws.append(["D", user["id"], "", "", "", ""])
    buf = BytesIO()
    wb.save(buf)
    xlsx = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    result = (
        await admin_client.post(
            "/api/admin/users/import", files={"file": ("users.xlsx", buf.getvalue(), xlsx)}
        )
    ).json()
    assert result["deleted"] == 1
    await _assert_history_detached(admin_client, session, user)
