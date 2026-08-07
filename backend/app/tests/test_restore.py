"""Restore-from-backup round trips over the real API + database.

The core contract under test: export -> mutate the database -> preview shows exactly the
mutations, inverted -> apply -> the database matches the export again, including the
*derived* item status (which only survives if the event history round-tripped).
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
async def world(admin_client):
    """Groups, users, items, a live loan, and anonymized history — a real small database."""
    group = (await admin_client.post("/api/admin/groups", json={"name": "Room 12"})).json()
    ada = (
        await admin_client.post("/api/admin/users", json={"name": "Ada", "group_id": group["id"]})
    ).json()
    grace = (await admin_client.post("/api/admin/users", json={"name": "Grace"})).json()
    eve = (await admin_client.post("/api/admin/users", json={"name": "Eve"})).json()
    calc_type = (
        await admin_client.post("/api/admin/item-types", json={"name": "Calculator"})
    ).json()
    calc1 = (
        await admin_client.post(
            "/api/admin/items",
            json={"item_type_id": calc_type["id"], "name": "Calc #1", "location": "Cabinet A"},
        )
    ).json()
    calc2 = (
        await admin_client.post(
            "/api/admin/items", json={"item_type_id": calc_type["id"], "name": "Calc #2"}
        )
    ).json()
    # A live loan: Ada holds Calc #1 at export time.
    resp = await admin_client.post(
        "/api/kiosk/checkout", json={"item_id": calc1["id"], "user_id": ada["id"]}
    )
    assert resp.status_code == 200
    # Anonymized history: Eve borrowed Calc #2, then was deleted -> her events stay,
    # user-less. The backup must round-trip those NULL-user events.
    for action in ("checkout", "checkin"):
        resp = await admin_client.post(
            f"/api/kiosk/{action}", json={"item_id": calc2["id"], "user_id": eve["id"]}
        )
        assert resp.status_code == 200
    assert (await admin_client.delete(f"/api/admin/users/{eve['id']}")).status_code == 204
    await admin_client.patch("/api/admin/settings", json={"label_width_mm": 40})
    return {
        "group": group,
        "ada": ada,
        "grace": grace,
        "type": calc_type,
        "calc1": calc1,
        "calc2": calc2,
    }


async def _export(admin_client) -> bytes:
    resp = await admin_client.get("/api/admin/database.xlsx")
    assert resp.status_code == 200
    return resp.content


def _upload(content: bytes) -> dict:
    return {"file": ("backup.xlsx", content, XLSX)}


def _entity(plan: dict, kind: str) -> dict:
    return next(e for e in plan["entities"] if e["kind"] == kind)


@pytest.mark.asyncio
async def test_restore_requires_admin(client):
    resp = await client.post("/api/admin/restore", files=_upload(b"x"))
    assert resp.status_code == 401


@pytest.mark.asyncio
async def test_round_trip_restores_data_history_and_derived_status(admin_client, world):
    backup = await _export(admin_client)

    # Mutate everything restore must undo.
    await admin_client.patch(f"/api/admin/users/{world['ada']['id']}", json={"name": "Ada L"})
    await admin_client.post("/api/admin/users", json={"name": "Zed"})
    assert (
        await admin_client.delete(f"/api/admin/items/{world['calc2']['id']}")
    ).status_code == 204
    resp = await admin_client.post(  # newer event the backup doesn't have
        "/api/kiosk/checkin", json={"item_id": world["calc1"]["id"], "user_id": world["ada"]["id"]}
    )
    assert resp.status_code == 200
    await admin_client.patch("/api/admin/settings", json={"label_width_mm": 50})

    # Preview: exactly the inverse of the mutations, and nothing is applied by it.
    plan = (await admin_client.post("/api/admin/restore", files=_upload(backup))).json()
    assert plan["errors"] == []
    assert plan["applied"] is False
    users = _entity(plan, "users")
    assert (users["create_count"], users["update_count"], users["delete_count"]) == (0, 1, 1)
    assert users["updates"][0]["fields"] == [{"field": "name", "old": "Ada L", "new": "Ada"}]
    assert users["deletes"][0]["label"].startswith("Zed")
    items = _entity(plan, "items")
    assert (items["create_count"], items["update_count"], items["delete_count"]) == (1, 0, 0)
    assert items["unchanged"] == 1
    assert plan["events_remove"] == 1  # the post-backup checkin
    assert plan["events_add"] == 3  # Calc #2's create + Eve's anonymized checkout/checkin
    assert plan["settings"] == [{"key": "label_width_mm", "old": "50.0", "new": "40.0"}]
    names = {u["name"] for u in (await admin_client.get("/api/admin/users")).json()}
    assert "Zed" in names and "Ada L" in names  # preview is read-only

    # Apply, then the database must match the backup.
    plan = (await admin_client.post("/api/admin/restore?apply=true", files=_upload(backup))).json()
    assert plan["errors"] == [] and plan["applied"] is True

    users_now = {u["name"]: u for u in (await admin_client.get("/api/admin/users")).json()}
    assert set(users_now) == {"Ada", "Grace"}
    assert users_now["Ada"]["barcode"] == world["ada"]["barcode"]
    items_now = {i["name"]: i for i in (await admin_client.get("/api/admin/items")).json()}
    assert set(items_now) == {"Calc #1", "Calc #2"}
    assert items_now["Calc #2"]["id"] == world["calc2"]["id"]
    assert items_now["Calc #2"]["barcode"] == world["calc2"]["barcode"]
    # The derived state came back: the loan exists again because the checkin was removed.
    assert items_now["Calc #1"]["status"] == "Checked out"
    settings = (await admin_client.get("/api/admin/settings")).json()
    assert settings["label_width_mm"] == 40.0

    # Restoring a database that already matches the backup is a no-op plan.
    plan = (await admin_client.post("/api/admin/restore", files=_upload(backup))).json()
    for entity in plan["entities"]:
        assert (entity["create_count"], entity["update_count"], entity["delete_count"]) == (
            0,
            0,
            0,
        ), entity["kind"]
    assert plan["events_add"] == 0 and plan["events_remove"] == 0
    assert plan["events_relink"] == 0
    assert plan["settings"] == []


@pytest.mark.asyncio
async def test_restore_is_all_or_nothing_on_bad_rows(admin_client, world):
    backup = await _export(admin_client)
    wb = load_workbook(BytesIO(backup))
    wb["users"].cell(row=2, column=2, value="not-a-uuid")  # corrupt one id
    buf = BytesIO()
    wb.save(buf)
    corrupted = buf.getvalue()

    before = (await admin_client.get("/api/admin/users")).json()
    for apply in ("false", "true"):
        plan = (
            await admin_client.post(f"/api/admin/restore?apply={apply}", files=_upload(corrupted))
        ).json()
        assert plan["applied"] is False
        assert any(e["sheet"] == "users" and e["row"] == 2 for e in plan["errors"])
        assert plan["entities"] == []  # no partial plan alongside errors
    assert (await admin_client.get("/api/admin/users")).json() == before


@pytest.mark.asyncio
async def test_not_a_backup_is_rejected(admin_client):
    plan = (await admin_client.post("/api/admin/restore", files=_upload(b"not xlsx"))).json()
    assert plan["errors"][0]["sheet"] == "workbook"
    # A random workbook (an import file, say) is missing the backup sheets.
    from app.tests.test_spreadsheet import _make_xlsx

    plan = (
        await admin_client.post(
            "/api/admin/restore", files=_upload(_make_xlsx(["action", "id"], []))
        )
    ).json()
    assert any(e["message"].startswith("Sheet is missing") for e in plan["errors"])


@pytest.mark.asyncio
async def test_old_backup_without_a_column_preserves_that_field(admin_client, world):
    backup = await _export(admin_client)
    wb = load_workbook(BytesIO(backup))
    ws = wb["items"]
    headers = [c.value for c in ws[1]]
    ws.delete_cols(headers.index("photo_url") + 1)  # simulate a pre-photo_url backup
    buf = BytesIO()
    wb.save(buf)
    old_format = buf.getvalue()

    await admin_client.patch(
        f"/api/admin/items/{world['calc1']['id']}", json={"photo_url": "http://pi/calc.jpg"}
    )
    plan = (
        await admin_client.post("/api/admin/restore?apply=true", files=_upload(old_format))
    ).json()
    assert plan["errors"] == [] and plan["applied"] is True
    items = {i["name"]: i for i in (await admin_client.get("/api/admin/items")).json()}
    assert items["Calc #1"]["photo_url"] == "http://pi/calc.jpg"  # absent column = untouched


@pytest.mark.asyncio
async def test_swapped_barcodes_restore_without_unique_collisions(admin_client, world):
    backup = await _export(admin_client)
    ada, grace = world["ada"], world["grace"]
    # Swap the two barcodes through a temp value (unique constraint blocks a direct swap).
    await admin_client.patch(f"/api/admin/users/{ada['id']}", json={"barcode": "U00000001"})
    await admin_client.patch(f"/api/admin/users/{grace['id']}", json={"barcode": ada["barcode"]})
    await admin_client.patch(f"/api/admin/users/{ada['id']}", json={"barcode": grace["barcode"]})

    plan = (await admin_client.post("/api/admin/restore?apply=true", files=_upload(backup))).json()
    assert plan["errors"] == [] and plan["applied"] is True
    users = {u["name"]: u for u in (await admin_client.get("/api/admin/users")).json()}
    assert users["Ada"]["barcode"] == ada["barcode"]
    assert users["Grace"]["barcode"] == grace["barcode"]
