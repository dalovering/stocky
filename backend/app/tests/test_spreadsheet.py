"""XLSX export/import round-trips for users and items, over the real API + database."""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _make_xlsx(headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _rows(content: bytes) -> list[dict]:
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in data[0]]
    return [dict(zip(headers, r, strict=False)) for r in data[1:]]


@pytest.fixture
async def seeded(admin_client):
    group = (await admin_client.post("/api/admin/groups", json={"name": "Room 12"})).json()
    user = (
        await admin_client.post("/api/admin/users", json={"name": "Ada", "group_id": group["id"]})
    ).json()
    item_type = (
        await admin_client.post("/api/admin/item-types", json={"name": "Calculator"})
    ).json()
    item = (
        await admin_client.post(
            "/api/admin/items",
            json={"item_type_id": item_type["id"], "name": "Calc #1", "location": "Cabinet A"},
        )
    ).json()
    return {"group": group, "user": user, "item_type": item_type, "item": item}


@pytest.mark.asyncio
async def test_export_users_has_action_column_and_data(admin_client, seeded):
    resp = await admin_client.get("/api/admin/users.xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX
    rows = _rows(resp.content)
    assert rows[0]["action"] is None  # action is blank on export
    assert rows[0]["name"] == "Ada"
    assert rows[0]["group"] == "Room 12"
    assert rows[0]["status"] == "Active"


@pytest.mark.asyncio
async def test_import_users_create_update_delete(admin_client, seeded):
    existing = seeded["user"]
    content = _make_xlsx(
        ["action", "id", "barcode", "name", "group", "status"],
        [
            ["C", "", "", "Grace Hopper", "Room 12", "Active"],  # create
            ["U", existing["id"], "", "Ada Lovelace", "", "Inactive"],  # update name + status
            ["", "", "", "ignored row", "", ""],  # blank action -> skipped
        ],
    )
    result = (
        await admin_client.post(
            "/api/admin/users/import", files={"file": ("users.xlsx", content, XLSX)}
        )
    ).json()
    assert result["created"] == 1
    assert result["updated"] == 1
    assert result["skipped"] == 1
    assert result["errors"] == []

    users = (await admin_client.get("/api/admin/users")).json()
    names = {u["name"]: u for u in users}
    assert "Grace Hopper" in names
    assert names["Ada Lovelace"]["status"] == "Inactive"


@pytest.mark.asyncio
async def test_import_items_round_trip(admin_client, seeded):
    # Export, then re-import with edits.
    export = (await admin_client.get("/api/admin/items.xlsx")).content
    rows = _rows(export)
    headers = [
        "action",
        "id",
        "barcode",
        "name",
        "item_type",
        "location",
        "condition",
        "needs_review",
    ]
    existing = rows[0]
    content = _make_xlsx(
        headers,
        [
            ["U", existing["id"], "", "Calc #1", "Calculator", "Shelf 5", "Fair", "FALSE"],
            ["C", "", "", "Calc #2", "Calculator", "Cabinet A", "New", "FALSE"],
        ],
    )
    result = (
        await admin_client.post(
            "/api/admin/items/import", files={"file": ("items.xlsx", content, XLSX)}
        )
    ).json()
    assert result["created"] == 1
    assert result["updated"] == 1

    items = (await admin_client.get("/api/admin/items")).json()
    by_name = {i["name"]: i for i in items}
    assert by_name["Calc #1"]["location"] == "Shelf 5"
    assert by_name["Calc #1"]["condition"] == "Fair"
    assert "Calc #2" in by_name


@pytest.mark.asyncio
async def test_import_reports_row_errors(admin_client, seeded):
    content = _make_xlsx(
        ["action", "id", "barcode", "name", "group", "status"],
        [
            ["C", "", "", "No Group User", "Nonexistent Room", "Active"],  # unknown group
            ["X", "", "", "Bad Action", "", ""],  # invalid action
        ],
    )
    result = (
        await admin_client.post(
            "/api/admin/users/import", files={"file": ("users.xlsx", content, XLSX)}
        )
    ).json()
    assert result["created"] == 0
    assert len(result["errors"]) == 2
    assert result["errors"][0]["row"] == 2  # first data row


@pytest.mark.asyncio
async def test_import_requires_admin(client):
    content = _make_xlsx(["action", "name"], [["C", "Ada"]])
    resp = await client.post(
        "/api/admin/users/import", files={"file": ("users.xlsx", content, XLSX)}
    )
    assert resp.status_code == 401


@pytest.mark.asyncio
async def test_export_history_xlsx(admin_client, seeded):
    # Build some history through the real API: a checkout and a checkin.
    user, item = seeded["user"], seeded["item"]
    await admin_client.post(
        "/api/kiosk/checkout", json={"item_id": item["id"], "user_id": user["id"]}
    )
    await admin_client.post(
        "/api/kiosk/checkin", json={"item_id": item["id"], "user_id": user["id"]}
    )

    resp = await admin_client.get("/api/admin/events.xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX
    rows = _rows(resp.content)
    assert list(rows[0].keys()) == [
        "id",
        "created_at",
        "event_type",
        "item",
        "item_barcode",
        "user",
        "user_barcode",
        "note",
    ]
    # create + checkout + checkin, most recent first, carrying names and barcodes.
    types = [r["event_type"] for r in rows]
    assert types == ["checkin", "checkout", "create"]
    assert rows[0]["item"] == "Calc #1"
    assert rows[0]["item_barcode"] == item["barcode"]
    assert rows[0]["user"] == "Ada"
    assert rows[0]["user_barcode"] == user["barcode"]
    # Timestamps come out as naive local wall time (openpyxl can't store tz-aware datetimes).
    assert rows[0]["created_at"].tzinfo is None


@pytest.mark.asyncio
async def test_export_history_xlsx_filters_by_user(admin_client, seeded):
    user, item = seeded["user"], seeded["item"]
    other = (await admin_client.post("/api/admin/users", json={"name": "Grace"})).json()
    await admin_client.post(
        "/api/kiosk/checkout", json={"item_id": item["id"], "user_id": user["id"]}
    )

    all_rows = _rows((await admin_client.get("/api/admin/events.xlsx")).content)
    narrowed = _rows(
        (await admin_client.get(f"/api/admin/events.xlsx?user_id={user['id']}")).content
    )
    assert len(narrowed) == 1  # just Ada's checkout; the item's create event has no user
    assert narrowed[0]["user"] == "Ada"
    assert len(all_rows) > len(narrowed)

    empty = _rows((await admin_client.get(f"/api/admin/events.xlsx?user_id={other['id']}")).content)
    assert empty == []


@pytest.mark.asyncio
async def test_export_history_requires_admin(client):
    assert (await client.get("/api/admin/events.xlsx")).status_code == 401


@pytest.mark.asyncio
async def test_full_database_export(admin_client, seeded):
    user, item = seeded["user"], seeded["item"]
    await admin_client.post(
        "/api/kiosk/checkout", json={"item_id": item["id"], "user_id": user["id"]}
    )

    resp = await admin_client.get("/api/admin/database.xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX

    wb = load_workbook(BytesIO(resp.content))
    assert wb.sheetnames == ["users", "groups", "item_types", "items", "history", "settings"]

    def sheet_rows(name: str) -> list[dict]:
        ws = wb[name]
        data = list(ws.iter_rows(values_only=True))
        headers = [str(h) for h in data[0]]
        return [dict(zip(headers, r, strict=False)) for r in data[1:]]

    users = sheet_rows("users")
    assert [u["name"] for u in users] == ["Ada"]
    assert users[0]["action"] is None  # import-compatible: blank action column

    groups = sheet_rows("groups")
    assert [g["name"] for g in groups] == ["Room 12"]

    types = sheet_rows("item_types")
    assert [t["name"] for t in types] == ["Calculator"]

    items = sheet_rows("items")
    assert [i["name"] for i in items] == ["Calc #1"]

    history = sheet_rows("history")
    assert [h["event_type"] for h in history] == ["checkout", "create"]

    # The settings sheet holds exactly the declared AppSettings keys — and above all, never the
    # admin password hash, which lives in the same settings table.
    settings_rows = sheet_rows("settings")
    keys = {r["key"] for r in settings_rows}
    assert "timezone" in keys
    assert "admin_password_hash" not in keys
    assert b"admin_password_hash" not in resp.content


@pytest.mark.asyncio
async def test_full_database_export_requires_admin(client):
    assert (await client.get("/api/admin/database.xlsx")).status_code == 401
